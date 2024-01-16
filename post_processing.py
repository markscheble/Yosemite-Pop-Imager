# Note: This software is Reserved Product developed by Planet Innovation
#
# Copyright (c) 2024, Planet Innovation
# 436 Elgar Rd, Box Hill, 3128, VIC, Australia
# Phone: +61 3 9945 7510
#
# The copyright to the computer program(s) herein is the property of
# Planet Innovation, Australia.
# The program(s) may be used and/or copied only with the written permission
# of Planet Innovation or in accordance with the terms and conditions
# stipulated in the agreement/contract under which the program(s) have been
# supplied.
#

import numpy as np
from scipy import stats as st
from image import Mask, Image
from IDS_Peak_Image_Acq import initialize_directory
from openpyxl import Workbook
from datetime import datetime
import math


def write_data(imgs, analysis_filename, numberofwells):
    """This function writes the image acquisition data to a csv"""
    # create excel workbook to write data
    max_char_count = 32767
    try:
        wb = Workbook()
        first = True
        # create iteration for each well
        for well_num in range(numberofwells):

            # if first input, name tab
            if first:
                ws = wb.active
                ws.title = f"Well {well_num}"
                first = False
            else:
                ws = wb.create_sheet(f"Well {well_num}")
            
            # write headers
            row, col = 1, 1
            ws.cell(row, col, "Time of Image Creation")
            ws.cell(row, col + 1, "Area")
            ws.cell(row, col + 2, "Mean")
            ws.cell(row, col + 3, "Stdev")
            ws.cell(row, col + 4, "Median")
            ws.cell(row, col + 5, "Minimum")
            ws.cell(row, col + 6, "Maximum")
            ws.cell(row, col + 7, "Mode")
            ws.cell(row, col + 8, "File Path")
            ws.cell(row, col + 9, "Pixel Int Array pt. 1")
            ws.cell(row, col + 10, "Pixel Int Array pt. 2")
            row += 1

            # create new row for each image taken
            for img in imgs:
                time = img.get_time()
                well = img.get_well(well_num)
                mean = well.get_mean()
                stdev = well.get_stdev()
                mode = well.get_mode()[0]
                minimum = well.get_min()
                maximum = well.get_max()
                area = well.get_area()
                pixel_int_array = well.get_pixel_int_array()
                median = well.get_median()
                filepath = img.get_file_path()
                ws.cell(row, col, time)
                ws.cell(row, col + 1, area)
                ws.cell(row, col + 2, mean)
                ws.cell(row, col + 3, stdev)
                ws.cell(row, col + 4, median)
                ws.cell(row, col + 5, minimum)
                ws.cell(row, col + 6, maximum)
                ws.cell(row, col + 7, mode)
                ws.cell(row, col + 8, filepath)

                #write pixel int array to file, if greater than max character cell limit - split into two
                str_pixel_int_array = str(pixel_int_array)
                length_pixel_arr = len(str_pixel_int_array)
                if length_pixel_arr > max_char_count:
                    ws.cell(row, col + 9, str_pixel_int_array[0:math.floor(length_pixel_arr/2)])
                    ws.cell(row, col + 10, str_pixel_int_array[math.floor(length_pixel_arr/2):])
                    ws.cell(row, col + 11, " ")
                else:
                    ws.cell(row, col + 9, str_pixel_int_array)
                    ws.cell(row, col + 10, " ")
                row += 1

        # save the workbook
        wb.save(analysis_filename)
        print("Workbook Saved")

    except Exception as e:
        print(f"Exception: {e}")

# WILL BE DELETED
def post_processing_test(logging = True, wells = 1, directory = "./Images/"):
    """This function tests the post processing and is called from gui.py after the images are taken"""

    # create the mask using the last image in array (when all well samples should be positive)
    mask_img = Mask("./Fluro_well2.png", num_wells=wells)

    # get the masks from the image
    mask_img.getMasks()

    print("-"*70)

    # print mask image data to console
    mask_img.print_all()

    # combine all individual masks in a singular mask for show
    mask_img.combineMasks(show=False)

    # save all created mask images
    folder = mask_img.saveImages(directory=directory)
    print(f"Saving in Directory: {folder}")

    # create Image class from image
    analysis_img = Image("./Fluro_well2.png", num_wells=wells)

    print("-"*70)

    # use the mask to call a binary and and isolate the individual well
    analysis_img.initialize_wells_from_mask(mask_img)

    print("-"*70)
    
    # analyze the image using array created above
    analysis_img.analyze_img()

    print("-"*70)

    # print well information from mask to console
    analysis_img.print_all()

    # save the images and masked images to directory
    # print(f"Saving in Directory: {directory}")
    # analysis_img.saveImages(directory=directory)
    
    if logging:
        # create filename for excel sheet
        analysis_filename = str("Yosemite_Area_Imager_Analysis.xlsx")

        # get directory
        directory = initialize_directory("./Yosemite_Area_Imager/")

        # get timestamp
        ts = datetime.now().strftime('./Yosemite_Area_Imager/%Y_%m_%d_%H-%M-%S_') 

        # write data to excel sheet
        write_data([analysis_img], directory+ts+analysis_filename, wells)



def post_processing(image_array, wells, logging, directory = "./Images/"):
    """This function will be called to post process the images taken from gui.py. """

    # assert length of image array is not less than or equal to 0
    assert (len(image_array) > 0), "No Images Taken"

    # create the mask using the last image in array (when all well samples should be positive)
    mask_img = Mask(image_array[-1], num_wells=wells)

    # get the masks from the image
    mask_img.getMasks()

    print("-"*70)

    # print mask image data to console
    mask_img.print_all()

    # combine all individual masks in a singular mask for show
    mask_img.combineMasks(show=False)

    # save all created mask images
    print(f"Saving in Directory: {directory}")
    mask_img.saveImages(directory=directory)

    # initialize analyzed image array
    analyzed_images = []

    for img in image_array:
        
        # create Image class from image
        analysis_img = Image(img, num_wells=wells)

        print("-"*70)

        # use the mask to call a binary and and isolate the individual well
        analysis_img.initialize_wells_from_mask(mask_img)

        print("-"*70)

        # analyze the image using array created above
        analysis_img.analyze_img()

        print("-"*70)

        # print well information from mask to console
        analysis_img.print_all()

        # save the images and masked images to directory
        # print(f"Saving in Directory: {directory}")
        # analysis_img.saveImages(directory=directory)

        # append the image to analyzed images
        analyzed_images.append(analysis_img)

    if logging:
        
        # create filename for excel sheet
        analysis_filename = str("Yosemite_Area_Imager_Analysis.xlsx")

        # get timestamp
        ts = datetime.now().strftime('%Y_%m_%d_%H-%M_') 

        # write data to excel sheet
        write_data(analyzed_images, directory + ts + analysis_filename, wells)



def post_processing_user(wells = 2, logging= True, image_dir = "./Preliminary images/Image 3, 10mm FL, LED mounted as close to well without masking, second LED held b, unknown imaging details.png", directory = "./Images/"):
    """This function is for user post-processing. """

    # create the mask using the last image in array (when all well samples should be positive)
    mask_img = Mask(image_dir, num_wells=wells)

    # get the masks from the image
    mask_img.getMasks()

    print("-"*70)

    # print mask image data to console
    mask_img.print_all()

    # combine all individual masks in a singular mask for show
    mask_img.combineMasks(show=False)

    # save all created mask images
    folder = mask_img.saveImages(directory=directory)
    print(f"Saving in Directory: {folder}")

        
    # create Image class from image
    analysis_img = Image(image_dir, num_wells=wells)

    print("-"*70)

    # use the mask to call a binary and and isolate the individual well
    analysis_img.initialize_wells_from_mask(mask_img)

    print("-"*70)

    # analyze the image using array created above
    analysis_img.analyze_img()

    print("-"*70)

    # print well information from mask to console
    analysis_img.print_all()

    # save the images and masked images to directory
    # print(f"Saving in Directory: {directory}")
    # folder = analysis_img.saveImages(directory=directory)

    if logging:
        
        # create filename for excel sheet
        analysis_filename = str("Yosemite_Area_Imager_Analysis.xlsx")

        # get timestamp
        ts = datetime.now().strftime('%Y_%m_%d_%H-%M_') 

        # write data to excel sheet
        write_data([analysis_img], folder+ts+analysis_filename, wells)




def post_processing_unit_test(image_array, wells, logging, directory = "./Images/"):
    """This function will be called to post process the images taken from gui.py. """

    # assert length of image array is not less than or equal to 0
    assert (len(image_array) > 0), "No Images Taken"

    # create the mask using the last image in array (when all well samples should be positive)
    mask_img = Mask(image_array[-1], num_wells=wells)

    # get the masks from the image
    mask_img.getMasks()

    print("-"*70)

    # print mask image data to console
    mask_img.print_all()

    # combine all individual masks in a singular mask for show
    mask_img.combineMasks(show=False)

    # save all created mask images
    folder = mask_img.saveImages(directory=directory)
    print(f"Saving in Directory: {folder}")

    # initialize analyzed image array
    analyzed_images = []

    for img in image_array:
        
        # create Image class from image
        analysis_img = Image(img, num_wells=wells)

        print("-"*70)

        # use the mask to call a binary and and isolate the individual well
        analysis_img.initialize_wells_from_mask(mask_img)

        print("-"*70)

        # analyze the image using array created above
        analysis_img.analyze_img()

        print("-"*70)

        # print well information from mask to console
        analysis_img.print_all()

        # save the images and masked images to directory
        # analysis_img.saveImages(directory=directory)

        # append the image to analyzed images
        analyzed_images.append(analysis_img)

    if logging:
        
        # create filename for excel sheet
        analysis_filename = str("Yosemite_Area_Imager_Analysis.xlsx")

        # get timestamp
        ts = datetime.now().strftime('%Y_%m_%d_%H-%M_') 

        # write data to excel sheet
        write_data(analyzed_images, folder+ts+analysis_filename, wells)
     
if __name__ == '__main__':
    # post_processing_user()

    single_well_image_arr  = ['./Preliminary images/Col LED, 10mm FL asp lens, 10FPS, 80ms Exp, 1.9 AGain.jpg', 
                              './Preliminary images/Col LED, 10mm FL asp lens, 10FPS, 95ms Exp, 3 AGain, Wired Separately.jpg',
                              './Preliminary images/Focused LED, 10mm FL asp lens, 10FPS, 95ms Exp, 3 AGain, Wired Separately.jpg',
                              './Preliminary images/Focused LED, 20mm FL asp lens 1in, 10FPS, 95ms Exp, 3 AGa, unmounted.jpg',
                              './Preliminary images/Focuses LED, 10mm FL asp lens, 10FPS, 80ms Exp, 1.9 AGa.jpg']
    
    two_well_image_arr = ['./Preliminary images/Col LED, 10mm FL asp lens, 5FPS,189ms Exp, 5 AGain, 30mm spot 2 Well.jpg'] 
                        #   './Preliminary images/Image 1 Col LED, 10mm FL, LED mounted as close to well without masking, second LED held b, unknown imaging details.png', 
                        #   './Preliminary images/Image 2, 10mm FL, LED mounted as close to well without masking, second LED held b, unknown imaging details.png', 
                        #   './Preliminary images/Image 3, 10mm FL, LED mounted as close to well without masking, second LED held b, unknown imaging details.png']
    
    five_well_image_arr = ["./20240112-Images/2024_01_11_14-23-44_0-1ng.png",
                            "./20240112-Images/2024_01_11_13-40-40_1ng.png",
                            "./20240112-Images/2024_01_11_12-30-03_5ng.png"]
    post_processing_unit_test(five_well_image_arr, 5, True)
    # post_processing_unit_test(two_well_image_arr, 2, True)



# Single Well
# Col LED, 10mm FL asp lens, 10FPS, 80ms Exp, 1.9 AGain
# Col LED, 10mm FL asp lens, 10FPS, 95ms Exp, 3 AGain, Wired Separately
# Focused LED, 10mm FL asp lens, 10FPS, 95ms Exp, 3 AGain, Wired Separately.jpg
# Focused LED, 20mm FL asp lens 1in, 10FPS, 95ms Exp, 3 AGa, unmounted
# Focuses LED, 10mm FL asp lens, 10FPS, 80ms Exp, 1.9 AGa
#

# Two Wells
# Col LED, 10mm FL asp lens, 5FPS,189ms Exp, 5 AGain, 30mm spot 2 Well
# Image 1 Col LED, 10mm FL, LED mounted as close to well without masking, second LED held b, unknown imaging details
# Image 2, 10mm FL, LED mounted as close to well without masking, second LED held b, unknown imaging details
# Image 3, 10mm FL, LED mounted as close to well without masking, second LED held b, unknown imaging details
# "./20240112-Images/2024_01_11_12-30-03_5ng.png"
# "./20240112-Images/2024_01_11_13-40-40_1ng.png"
# "./20240112-Images/2024_01_11_14-23-44_0-1ng.png"